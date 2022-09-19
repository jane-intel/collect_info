#!/usr/bin/env python3

import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils import shell, get_all_xmls

if __name__ == '__main__':
    ops_which_stop_dynamic_rank = set()
    ops_which_stop_dynamic_type = set()
    ops_with_dynamic_output_rank = defaultdict(int)
    ops_with_dynamic_element_type = defaultdict(int)

    all_models = get_all_xmls(sys.argv)

    wb = Workbook()
    tsheet = wb.active
    tsheet.title = "Models tested"
    tsheet.append(["Path to model", "Test status"])
    tsheet.column_dimensions[get_column_letter(1)].width = 50

    for i, model in enumerate(all_models):
        print("Results for:", model)
        return_code, out, err = shell([sys.executable, " per_model_test.py ", str(model)])
        if return_code != 0 or len(err) != 0:
            print("Some error happened:")
            print(err)
            tsheet.append([str(model), err.strip()])
            continue
        print(out)
        for line in out.split("\n"):
            if not line:
                continue
            if 'DYN_OUT_RANK' in line:
                result = line.replace('DYN_OUT_RANK', '').strip().split(" ")
                assert len(result) == 2, result
                ops_with_dynamic_output_rank[result[0]] += int(result[1])
            if 'STOP_DYN_RANK' in line:
                result = line.replace('STOP_DYN_RANK', '').strip().split(" ")
                assert len(result) == 1, result
                ops_which_stop_dynamic_rank.add(result[0])
            if 'DYN_OUT_TYPE' in line:
                result = line.replace('DYN_OUT_TYPE', '').strip().split(" ")
                assert len(result) == 2, result
                ops_with_dynamic_element_type[result[0]] += int(result[1])
            if 'STOP_DYN_TYPE' in line:
                result = line.replace('STOP_DYN_TYPE', '').strip().split(" ")
                assert len(result) == 1, result
                ops_which_stop_dynamic_type.add(result[0])
        tsheet.append([str(model), "Ok"])
        print("=" * 100)

    print("Statistics per all models:")
    tsheet = wb.create_sheet("Ops on DYN RANK path")
    tsheet.append(["Op name", "Number of ops"])
    print("Operations which propagated dynamic rank through:")
    for name, num_nodes in ops_with_dynamic_output_rank.items():
        print(name, num_nodes)
        tsheet.append([name, num_nodes])
    tsheet.auto_filter.ref = 'A1:' + get_column_letter(tsheet.max_column) + str(tsheet.max_row)
    tsheet.column_dimensions[get_column_letter(1)].width = 50
    tsheet.column_dimensions[get_column_letter(2)].width = 10

    tsheet = wb.create_sheet("Ops stopped DYN RANK")
    tsheet.append(["Op name"])
    print("Operations which stopped dynamic output rank from propagating further:")
    for name in ops_which_stop_dynamic_rank:
        print(name)
        tsheet.append([name])
    tsheet.auto_filter.ref = 'A1:' + get_column_letter(tsheet.max_column) + str(tsheet.max_row)
    tsheet.column_dimensions[get_column_letter(1)].width = 50
    tsheet.column_dimensions[get_column_letter(2)].width = 10

    tsheet = wb.create_sheet("Ops on DYN TYPE path")
    tsheet.append(["Op name", "Number of ops"])
    print("Operations which propagated dynamic type through:")
    for name, num_nodes in ops_with_dynamic_element_type.items():
        print(name, num_nodes)
        tsheet.append([name, num_nodes])
    tsheet.auto_filter.ref = 'A1:' + get_column_letter(tsheet.max_column) + str(tsheet.max_row)
    tsheet.column_dimensions[get_column_letter(1)].width = 50
    tsheet.column_dimensions[get_column_letter(2)].width = 10

    tsheet = wb.create_sheet("Ops stopped DYN TYPE")
    tsheet.append(["Op name"])
    print("Operations which stopped dynamic output type from propagating further:")
    for name in ops_which_stop_dynamic_type:
        print(name)
        tsheet.append([name])
    tsheet.auto_filter.ref = 'A1:' + get_column_letter(tsheet.max_column) + str(tsheet.max_row)
    tsheet.column_dimensions[get_column_letter(1)].width = 50
    tsheet.column_dimensions[get_column_letter(2)].width = 10

    wb.save("wb.xlsx")
